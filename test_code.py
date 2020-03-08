import openpyxl
import os
from openpyxl.utils import get_column_letter,column_index_from_string

print(os.getcwd())

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

sheet3 = wb['Sheet1']

print(sheet3)
print(sheet3.title)

anotherSheet = wb.active
print(anotherSheet)


a = sheet3['B2']
print(a.value)

print('Row ' + str(a.row) + ', Column ' + str(a.column) + ' is ' + a.value)
print('C1 value is: ' + str(sheet3['C1'].value))

print(sheet3.cell(row=2, column=2).value)

for i in range(1,8,2):
    print(i, sheet3.cell(row=i, column=2).value)

print(sheet3.max_row)
print(sheet3.max_column)

print(get_column_letter(sheet3.max_column))
print('column_index_from_string: ' + str(column_index_from_string('A')))
